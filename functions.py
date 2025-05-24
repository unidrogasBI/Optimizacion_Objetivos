from pydrive2.auth import GoogleAuth
from pydrive2.drive import GoogleDrive
import customtkinter as ctk
import pandas as pd
from datetime import datetime
from tkinter import filedialog
import tkinter.messagebox as mbox
from openpyxl import load_workbook
import sys
import os

def resource_path(relative_path):
    """Obtiene la ruta absoluta a un recurso, funciona con PyInstaller o en desarrollo."""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def Boton_Subir():
    try:
        gauth = GoogleAuth()

        # Asegúrate de que este JSON exista en el mismodirectorio
        gauth.settings['service_config'] = {
            "client_json_file_path": resource_path("service_account.json"),
            "client_user_email":"undobjetivos@organic-acronym-457612-b3.iam.gserviceaccount.com"
        }
        try:
            gauth.ServiceAuth()
            drive = GoogleDrive(gauth)

            # Aquí podrías continuar con la subida de archivos, por ejemplo:
            file = drive.CreateFile({'title': 'objetivos.xlsx'})
            file.SetContentFile('objetivos.xlsx')  # Cambia por el nombre correcto
            file.Upload()

            print("✅ Archivo subido correctamente a Google Drive.")

        except Exception as e:
            print("❌ Error al intentar subir el archivo a GoogleDrive", e)
    except:
        print("Problema con el programa")

def cargar_datos_excel(ruta_excel):
    df = pd.read_excel(resource_path(ruta_excel), skiprows=7)
    df.columns = df.columns.str.strip()
    df = df.dropna(subset=["Región", "Zona Supervisión"])
    return df

def obtener_regiones(df):
    return df["Región"].dropna().unique().tolist()

def obtener_zonas_por_region(df, region):
    zonas = df[df["Región"] == region]["Zona Supervisión"].dropna().unique().tolist()
    return zonas

def cargar_archivo():
    archivo = filedialog.askopenfilename(title="Selecciona un archivo Excel", filetypes=[("Archivos Excel", "*.xlsx;*.xls")])
    
    if archivo:
        df = pd.read_excel(archivo)
        return df 
    else:
        return None

id_archivo = "1nYcTk9wJh7oCfh7u_W3EXdnIwaeSEDt0"

def agregar_filas_a_drive(df_nuevo):
    gauth = GoogleAuth()
    gauth.settings['service_config'] = {
        "client_json_file_path": resource_path("service_account.json"),
        "client_user_email": "undobjetivos@organic-acronym-457612-b3.iam.gserviceaccount.com"
    }
    try:
        gauth.ServiceAuth()
        drive = GoogleDrive(gauth)
        # Aquí podrías continuar con la subida de archivos, por ejemplo:
        archivo_drive = drive.CreateFile({'id': id_archivo})
        archivo_drive.GetContentFile(resource_path("temp.xlsx"))  # Cambia por el nombre correcto
        libro = load_workbook(resource_path("temp.xlsx"))
        hoja = libro.active
        filas_existentes = hoja.max_row
        for fila in df_nuevo.itertuples(index=False):
            hoja.append(fila)
        libro.save(resource_path("temp.xlsx"))
        # Subir archivo actualizado
        archivo_drive.SetContentFile(resource_path("temp.xlsx"))
        archivo_drive.Upload()
        print("✅ Archivo subido correctamente a Google Drive.")
    except Exception as e:
        print("❌ Error durante autenticación o subida:", e)

def guardar_copia_seguridad(local_path, nombre_archivo, folder_id, drive):
    try:
        print(f"Subiendo archivo: {local_path} como '{nombre_archivo}'")
        archivo_drive = drive.CreateFile({
            'title': nombre_archivo,
            'parents': [{'id': folder_id}]
        })
        archivo_drive.SetContentFile(local_path)
        archivo_drive.Upload()
        print("✅ Archivo subido correctamente a Drive.")
        return archivo_drive['id']
    except Exception as e:
        print("❌ Error al subir el archivo:", e)
        return None

def guardar_excel_local(df, ruta_archivo):
    df.to_excel(resource_path("Data/Temp/" + ruta_archivo), index=False)
    print(f"✅ Excel guardado en {ruta_archivo}")
    return ruta_archivo